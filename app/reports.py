"""Reports and export routes."""

from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for
from flask_login import login_required, current_user
from app.models import Expense
from datetime import datetime, timedelta
from io import BytesIO
import csv

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')


@reports_bp.route('/')
@login_required
def view_reports():
    """View reports page."""
    now = datetime.now()
    current_year = now.year
    current_month = now.month
    
    # Get monthly data
    month_start = datetime(current_year, current_month, 1).date()
    if current_month == 12:
        month_end = datetime(current_year + 1, 1, 1).date() - timedelta(days=1)
    else:
        month_end = datetime(current_year, current_month + 1, 1).date() - timedelta(days=1)
    
    monthly_expenses = Expense.query.filter(
        Expense.user_id == current_user.id,
        Expense.date >= month_start,
        Expense.date <= month_end
    ).all()
    
    monthly_total = sum(e.amount for e in monthly_expenses)
    monthly_by_category = {}
    for expense in monthly_expenses:
        monthly_by_category[expense.category] = monthly_by_category.get(
            expense.category, 0
        ) + expense.amount
    
    # Get yearly data
    year_start = datetime(current_year, 1, 1).date()
    year_end = datetime(current_year, 12, 31).date()
    
    yearly_expenses = Expense.query.filter(
        Expense.user_id == current_user.id,
        Expense.date >= year_start,
        Expense.date <= year_end
    ).all()
    
    yearly_total = sum(e.amount for e in yearly_expenses)
    yearly_by_category = {}
    for expense in yearly_expenses:
        yearly_by_category[expense.category] = yearly_by_category.get(
            expense.category, 0
        ) + expense.amount
    
    # Monthly breakdown for the year
    monthly_totals = {}
    for month in range(1, 13):
        month_start = datetime(current_year, month, 1).date()
        if month == 12:
            month_end = datetime(current_year + 1, 1, 1).date() - timedelta(days=1)
        else:
            month_end = datetime(current_year, month + 1, 1).date() - timedelta(days=1)
        
        month_total = sum(
            e.amount for e in yearly_expenses
            if month_start <= e.date <= month_end
        )
        monthly_totals[datetime(current_year, month, 1).strftime('%B')] = month_total
    
    return render_template('reports.html',
                           current_month=current_month,
                           current_year=current_year,
                           monthly_total=monthly_total,
                           monthly_by_category=monthly_by_category,
                           yearly_total=yearly_total,
                           yearly_by_category=yearly_by_category,
                           monthly_totals=monthly_totals)


@reports_bp.route('/export/pdf')
@login_required
def export_pdf():
    """Export expenses to PDF."""
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Get expenses
    query = Expense.query.filter_by(user_id=current_user.id)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Expense.date >= date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Expense.date <= date_to_obj)
        except ValueError:
            pass
    
    expenses = query.order_by(Expense.date.desc()).all()
    
    # Create PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1f77b4'),
        spaceAfter=30,
        alignment=1
    )
    elements.append(Paragraph('Expense Report', title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Report info
    if date_from and date_to:
        info_text = f'Period: {date_from} to {date_to}'
    elif date_from:
        info_text = f'From: {date_from}'
    elif date_to:
        info_text = f'To: {date_to}'
    else:
        info_text = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    
    elements.append(Paragraph(info_text, styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))
    
    # Table data
    table_data = [['Date', 'Category', 'Description', 'Amount']]
    total = 0
    
    for expense in expenses:
        table_data.append([
            expense.date.strftime('%Y-%m-%d'),
            expense.category,
            expense.description or '-',
            f'${expense.amount:.2f}'
        ])
        total += expense.amount
    
    table_data.append(['', '', 'TOTAL', f'${total:.2f}'])
    
    # Create table
    table = Table(table_data, colWidths=[1.2*inch, 1.5*inch, 2*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d4e6f1')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f'expense_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)


@reports_bp.route('/export/excel')
@login_required
def export_excel():
    """Export expenses to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Get expenses
    query = Expense.query.filter_by(user_id=current_user.id)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Expense.date >= date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Expense.date <= date_to_obj)
        except ValueError:
            pass
    
    expenses = query.order_by(Expense.date.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Expenses'
    
    # Style definitions
    header_fill = PatternFill(start_color='1f77b4', end_color='1f77b4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Date', 'Category', 'Description', 'Amount']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Data
    total = 0
    for row, expense in enumerate(expenses, 2):
        ws.cell(row=row, column=1).value = expense.date
        ws.cell(row=row, column=2).value = expense.category
        ws.cell(row=row, column=3).value = expense.description or ''
        ws.cell(row=row, column=4).value = expense.amount
        total += expense.amount
        
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col == 4:
                cell.number_format = '$#,##0.00'
    
    # Total row
    total_row = len(expenses) + 2
    ws.cell(row=total_row, column=3).value = 'TOTAL'
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    ws.cell(row=total_row, column=4).value = total
    ws.cell(row=total_row, column=4).font = Font(bold=True)
    ws.cell(row=total_row, column=4).number_format = '$#,##0.00'
    ws.cell(row=total_row, column=4).fill = PatternFill(start_color='d4e6f1', end_color='d4e6f1', fill_type='solid')
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f'expense_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@reports_bp.route('/export/csv')
@login_required
def export_csv():
    """Export expenses to CSV."""
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Get expenses
    query = Expense.query.filter_by(user_id=current_user.id)
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Expense.date >= date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Expense.date <= date_to_obj)
        except ValueError:
            pass
    
    expenses = query.order_by(Expense.date.desc()).all()
    
    # Create CSV in memory
    buffer = BytesIO()
    wrapper = buffer.write
    
    writer = csv.writer(wrapper)
    writer.writerow(['Date', 'Category', 'Description', 'Amount'])
    
    total = 0
    for expense in expenses:
        writer.writerow([expense.date, expense.category, expense.description or '', expense.amount])
        total += expense.amount
    
    writer.writerow(['', '', 'TOTAL', total])
    
    buffer.seek(0)
    
    filename = f'expense_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    return send_file(buffer, mimetype='text/csv', as_attachment=True, download_name=filename)
